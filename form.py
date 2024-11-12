from datetime import datetime
from openpyxl import load_workbook

from model import Model
from sharepoint import Sharepoint


class Form:
    def __init__(self, access_token, spdir_parent, excel_name):
        self.access_token = access_token
        self.excel_name = excel_name
        self.sp_parent = Sharepoint(access_token, spdir_parent)
        self.sp_item_id = self.get_item_id()
        self.rows = []

    # Get the item id of the excel file in the parent directory in SharePoint
    def get_item_id(self):
        try:
            items = self.sp_parent.get_children()
            for item in items:
                if item["name"] == self.excel_name:
                    return item["id"]
            # Raise an error if the excel file is not found
            raise FileNotFoundError(
                f"Excel file '{self.excel_name}' not found in the SharePoint folder."
            )
        except Exception as e:
            print(f"Error getting item id of '{self.excel_name}': {e}")
            raise

    # Download the excel file from the parent directory in SharePoint
    def download(self):
        try:
            self.sp_parent.download(self.excel_name, self.sp_item_id)
            print(f"Downloaded '{self.excel_name}' successfully.")
        except Exception as e:
            print(f"Error downloading '{self.excel_name}': {e}")
            raise

    # Upload the updated excel file back to the parent directory in SharePoint
    def upload(self):
        try:
            self.sp_parent.upload(self.excel_name)
            print(f"Uploaded '{self.excel_name}' successfully.")
        except Exception as e:
            print(f"Error uploading '{self.excel_name}': {e}")
            raise

    # Load the workbook and read the data from the excel file to rows
    def read(self):
        try:
            wb = load_workbook(self.excel_name)
            sheet = wb.active
            for row_idx in range(6, sheet.max_row + 1):
                row = Model(sheet, row_idx)
                # Append data to list if already processed
                if row.send == "Gönder." and not row.share_status == "Gönderildi.":
                    self.rows.append((row_idx, row))
            print(f"Read '{self.excel_name}' successfully.")
            return self.rows
        except Exception as e:
            print(f"Error reading '{self.excel_name}': {e}")
            raise
        finally:
            wb.close()

    # Write the updated rows to the excel file
    def write(self):
        try:
            wb = load_workbook(self.excel_name)
            sheet = wb.active
            for row_idx, row in self.rows:
                if row.share_status == "Gönderildi.":
                    sheet[f"V{row_idx}"] = row.share_url
                    sheet[f"W{row_idx}"] = row.share_date
                    sheet[f"X{row_idx}"] = row.share_status
            wb.save(self.excel_name)
            print(f"Updated '{self.excel_name}' successfully.")
        except Exception as e:
            print(f"Error writing to '{self.excel_name}': {e}")
            raise
        finally:
            wb.close()

    # Lock the excel file to prevent further editing
    def lock(self):
        try:
            self.sp_parent.checkout(self.sp_item_id)
            print(f"Locked '{self.excel_name}' successfully.")
        except Exception as e:
            print(f"Error locking '{self.excel_name}': {e}")
            raise

    # Unlock the excel file to allow editing
    def unlock(self, discard=False):
        try:
            if discard:
                self.sp_parent.discard_checkout(self.sp_item_id)
            else:
                self.sp_parent.checkin(self.sp_item_id)
            print(f"Unlocked '{self.excel_name}' successfully.")
        except Exception as e:
            print(f"Error unlocking '{self.excel_name}': {e}")
            raise
